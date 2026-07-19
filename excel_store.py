# -*- coding: utf-8 -*-
"""按「难度+关卡」持久化通关时间（JSON 主存 + Excel 同步）。"""
from __future__ import annotations

import json
import os
import shutil
import threading
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path
from typing import Deque, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_PER_STAGE = 10
HISTORY_SHEET = "明细"
SUMMARY_SHEET = "汇总"
DIFFICULTIES = ("普通", "噩梦", "地狱", "折磨", "未知")


def _parse_dt(value) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    if not text:
        return datetime.now()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(text[:26], fmt)
        except Exception:
            continue
    try:
        return datetime.fromisoformat(text)
    except Exception:
        return datetime.now()


def normalize_difficulty(name: str) -> str:
    text = str(name or "").strip()
    if not text:
        return "未知"
    for d in ("折磨", "地狱", "噩梦", "普通"):
        if d in text:
            return d
    if text in DIFFICULTIES:
        return text
    return "未知"


def make_stage_key(difficulty: str, stage: str) -> str:
    """存储主键：难度|关卡，例如 折磨|2-9"""
    return f"{normalize_difficulty(difficulty)}|{str(stage or '').strip()}"


def split_stage_key(key: str) -> Tuple[str, str]:
    key = str(key or "").strip()
    if "|" in key:
        left, right = key.split("|", 1)
        return normalize_difficulty(left), right.strip()
    # 兼容旧数据：仅关卡号
    return "未知", key


class ExcelStore:
    """
    主键为「难度|关卡」。
    每键保留最近 max_per_stage 次秒数并算平均。
    """

    def __init__(
        self,
        excel_path: Path,
        max_per_stage: int = MAX_PER_STAGE,
        json_path: Optional[Path] = None,
    ):
        self.path = Path(excel_path)
        self.json_path = Path(json_path) if json_path else self.path.with_suffix(".json")
        self.max_per_stage = max(1, int(max_per_stage))
        self._lock = threading.RLock()
        # key -> deque of (seconds, recorded_at, notice_time)
        self._data: Dict[str, Deque[Tuple[int, datetime, str]]] = defaultdict(
            lambda: deque(maxlen=self.max_per_stage)
        )
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._last_excel_error = ""
        self._last_excel_path = self.path
        self._load_persistent()
        try:
            self._write_json()
            ok, path, err = self._write_workbook_safe()
            if ok:
                self._last_excel_path = path
            else:
                self._last_excel_error = err
        except Exception:
            pass

    def _new_deque(self) -> Deque[Tuple[int, datetime, str]]:
        return deque(maxlen=self.max_per_stage)

    def _load_persistent(self) -> None:
        if self._load_json():
            return
        if self._load_excel():
            try:
                self._write_json()
            except Exception:
                pass
            return
        if not self.path.exists() and not self.json_path.exists():
            try:
                self._write_workbook()
                self._write_json()
            except Exception:
                pass

    def _load_json(self) -> bool:
        if not self.json_path.exists():
            return False
        try:
            raw = json.loads(self.json_path.read_text(encoding="utf-8"))
            stages = raw.get("stages") if isinstance(raw, dict) else None
            if not isinstance(stages, dict):
                return False
            loaded: Dict[str, Deque[Tuple[int, datetime, str]]] = {}
            for key, items in stages.items():
                key = str(key).strip()
                if not key or not isinstance(items, list):
                    continue
                # 旧版 key 可能只是 "2-9"，统一成 未知|2-9
                if "|" not in key:
                    key = make_stage_key("未知", key)
                else:
                    d, s = split_stage_key(key)
                    key = make_stage_key(d, s)
                dq = self._new_deque()
                for it in items:
                    if not isinstance(it, dict):
                        continue
                    try:
                        sec = int(it.get("seconds"))
                    except Exception:
                        continue
                    dt = _parse_dt(it.get("recorded_at"))
                    notice = str(it.get("notice_time") or "")
                    dq.append((sec, dt, notice))
                if dq:
                    loaded[key] = dq
            self._data = defaultdict(self._new_deque, loaded)
            return True
        except Exception:
            return False

    def _load_excel(self) -> bool:
        if not self.path.exists():
            return False
        try:
            wb = load_workbook(self.path, data_only=False)
            if HISTORY_SHEET not in wb.sheetnames:
                wb.close()
                return False
            ws = wb[HISTORY_SHEET]
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            # 兼容：关卡 | 难度+关卡 两列
            has_diff_col = "难度" in headers
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                if has_diff_col:
                    difficulty = normalize_difficulty(row[0] if len(row) > 0 else "")
                    stage = str(row[1] if len(row) > 1 else "").strip()
                    sec_i, dt_i, notice_i = 2, 3, 4
                else:
                    difficulty = "未知"
                    stage = str(row[0] if len(row) > 0 else "").strip()
                    sec_i, dt_i, notice_i = 1, 2, 3
                if not stage:
                    continue
                try:
                    sec = int(row[sec_i])
                except Exception:
                    continue
                dt = _parse_dt(row[dt_i] if len(row) > dt_i else None)
                notice = str(row[notice_i] or "") if len(row) > notice_i else ""
                rows.append((make_stage_key(difficulty, stage), sec, dt, notice))
            wb.close()
            rows.sort(key=lambda x: x[2])
            loaded: Dict[str, Deque[Tuple[int, datetime, str]]] = {}
            for key, sec, dt, notice in rows:
                if key not in loaded:
                    loaded[key] = self._new_deque()
                loaded[key].append((sec, dt, notice))
            if not loaded:
                return False
            self._data = defaultdict(self._new_deque, loaded)
            return True
        except Exception:
            return False

    def _write_json(self) -> None:
        payload = {
            "version": 2,
            "max_per_stage": self.max_per_stage,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "stages": {},
        }
        for key, items in self._data.items():
            difficulty, stage = split_stage_key(key)
            payload["stages"][key] = [
                {
                    "difficulty": difficulty,
                    "stage": stage,
                    "seconds": sec,
                    "recorded_at": dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "notice_time": notice,
                }
                for sec, dt, notice in items
            ]
        tmp = self.json_path.with_suffix(self.json_path.suffix + ".tmp")
        tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(self.json_path)

    @staticmethod
    def make_dedupe_key(
        difficulty: str,
        stage: str,
        clear_seconds: int,
        notice_time: str = "",
        recorded_at: Optional[datetime] = None,
    ) -> str:
        """去重键必须带日期。

        游戏通知时钟只有 HH:MM，若永久用「关卡+秒数+时钟」去重，
        会导致同一天之后同秒数通关再也写不进去，「最近记录时间」卡死。
        列表刷新重复：同一天内相同 toast 仍会跳过。
        """
        d = normalize_difficulty(difficulty)
        s = str(stage or "").strip()
        notice = str(notice_time or "").strip()
        sec = int(clear_seconds)
        when = recorded_at or datetime.now()
        day = when.strftime("%Y-%m-%d")
        if notice:
            return f"{d}|{s}|{sec}|{day}|{notice}"
        # 无通知时钟：按本地分钟桶，避免永久锁死同秒数
        return f"{d}|{s}|{sec}|{day}|{when.strftime('%H:%M')}"

    def add_clear(
        self,
        stage: str,
        clear_seconds: int,
        notice_time: str = "",
        recorded_at: Optional[datetime] = None,
        difficulty: str = "未知",
        *,
        allow_duplicate: bool = False,
    ) -> dict:
        stage = str(stage or "").strip()
        if not stage:
            raise ValueError("empty stage")
        sec = int(clear_seconds)
        if sec < 0 or sec > 86400:
            raise ValueError(f"invalid seconds: {sec}")
        dt = recorded_at or datetime.now()
        notice = str(notice_time or "").strip()
        difficulty = normalize_difficulty(difficulty)
        key = make_stage_key(difficulty, stage)
        dedupe = self.make_dedupe_key(difficulty, stage, sec, notice, dt)

        with self._lock:
            if not allow_duplicate:
                for old_sec, old_dt, old_notice in self._data.get(key, []):
                    old_key = self.make_dedupe_key(
                        difficulty, stage, old_sec, old_notice, old_dt
                    )
                    if old_key == dedupe:
                        snap = self.stage_snapshot(key)
                        snap["skipped"] = True
                        snap["dedupeKey"] = dedupe
                        snap["skipReason"] = "same-day-toast"
                        return snap
            self._data[key].append((sec, dt, notice))
            self._persist_all()
            snap = self.stage_snapshot(key)
            snap["skipped"] = False
            snap["dedupeKey"] = dedupe
            return snap

    def _persist_all(self) -> None:
        """JSON 必写；Excel 失败不回滚 JSON（占用时写备用文件）。"""
        self._write_json()
        ok, path, err = self._write_workbook_safe()
        if not ok:
            # 不 raise：否则调用方以为整次通关失败；JSON 已是真相
            self._last_excel_error = err
        else:
            self._last_excel_error = ""
            self._last_excel_path = path

    @property
    def last_excel_error(self) -> str:
        return str(getattr(self, "_last_excel_error", "") or "")

    @property
    def last_excel_path(self) -> Path:
        p = getattr(self, "_last_excel_path", None)
        return Path(p) if p else self.path

    def stage_snapshot(self, key_or_stage: str, difficulty: str = "") -> dict:
        if difficulty:
            key = make_stage_key(difficulty, key_or_stage)
        elif "|" in str(key_or_stage):
            key = str(key_or_stage)
        else:
            key = make_stage_key("未知", key_or_stage)
        with self._lock:
            items = list(self._data.get(key, []))
        d, s = split_stage_key(key)
        # items 按时间从旧到新；展示用「最近」取时间最大的一条
        if items:
            newest = max(items, key=lambda x: x[1] if isinstance(x[1], datetime) else _parse_dt(x[1]))
            secs_old_to_new = [x[0] for x in items]
            last_sec = newest[0]
            last_at = newest[1]
            if not isinstance(last_at, datetime):
                last_at = _parse_dt(last_at)
            last_notice = newest[2]
        else:
            secs_old_to_new = []
            last_sec = None
            last_at = None
            last_notice = ""
        avg = (
            round(sum(secs_old_to_new) / len(secs_old_to_new), 2) if secs_old_to_new else 0.0
        )
        return {
            "key": key,
            "difficulty": d,
            "stage": s,
            "count": len(secs_old_to_new),
            "seconds": secs_old_to_new,
            "average": avg,
            "last_seconds": last_sec,
            "last_at": last_at.strftime("%Y-%m-%d %H:%M:%S") if last_at else "",
            "notice_time": last_notice or "",
            "display": f"{d} {s}" if d != "未知" else s,
        }

    def all_snapshots(self) -> List[dict]:
        with self._lock:
            keys = sorted(self._data.keys(), key=self._key_sort)
        return [self.stage_snapshot(k) for k in keys]

    def record_count(self) -> int:
        with self._lock:
            return sum(len(v) for v in self._data.values())

    @staticmethod
    def _key_sort(key: str):
        d, s = split_stage_key(key)
        order = {"普通": 0, "噩梦": 1, "地狱": 2, "折磨": 3, "未知": 9}
        parts = s.replace("－", "-").split("-")
        try:
            a, b = int(parts[0]), int(parts[1]) if len(parts) > 1 else 0
        except Exception:
            a, b = 9999, 9999
        return (order.get(d, 8), a, b)

    def _build_workbook(self) -> Workbook:
        """汇总布局与 v1.1.2 一致：难度|关卡|第1~10次|样本数|平均|最近记录时间（无通知钟）。"""
        wb = Workbook()
        ws_sum = wb.active
        ws_sum.title = SUMMARY_SHEET
        headers = (
            ["难度", "关卡"]
            + [f"第{i}次(秒)" for i in range(1, self.max_per_stage + 1)]
            + ["样本数", "平均通关(秒)", "最近记录时间"]
        )
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws_sum.cell(1, col, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        keys = sorted(self._data.keys(), key=self._key_sort)
        avg_fill = PatternFill("solid", fgColor="FFF2CC")
        for r, key in enumerate(keys, 2):
            difficulty, stage = split_stage_key(key)
            items = list(self._data[key])
            # 旧 → 新（与 1.1.2 相同）
            secs = [x[0] for x in items]
            avg = round(sum(secs) / len(secs), 2) if secs else ""
            if items:
                newest = max(
                    items,
                    key=lambda x: x[1] if isinstance(x[1], datetime) else _parse_dt(x[1]),
                )
                last_dt = newest[1] if isinstance(newest[1], datetime) else _parse_dt(newest[1])
            else:
                last_dt = None

            ws_sum.cell(r, 1, difficulty)
            ws_sum.cell(r, 2, stage)
            for i in range(self.max_per_stage):
                ws_sum.cell(r, 3 + i, secs[i] if i < len(secs) else "")
            ws_sum.cell(r, 3 + self.max_per_stage, len(secs))
            avg_cell = ws_sum.cell(r, 4 + self.max_per_stage, avg)
            avg_cell.fill = avg_fill
            avg_cell.font = Font(bold=True)
            time_cell = ws_sum.cell(
                r,
                5 + self.max_per_stage,
                last_dt.strftime("%Y-%m-%d %H:%M:%S") if last_dt else "",
            )
            time_cell.alignment = Alignment(horizontal="center")

        for col in range(1, len(headers) + 1):
            ws_sum.column_dimensions[get_column_letter(col)].width = 14
        ws_sum.column_dimensions["A"].width = 10
        ws_sum.column_dimensions["B"].width = 10
        ws_sum.column_dimensions[get_column_letter(5 + self.max_per_stage)].width = 20

        ws_hist = wb.create_sheet(HISTORY_SHEET)
        hist_headers = ["难度", "关卡", "通关秒数", "记录时间", "序号(该关内)"]
        for col, h in enumerate(hist_headers, 1):
            cell = ws_hist.cell(1, col, h)
            cell.fill = header_fill
            cell.font = header_font
        row_i = 2
        for key in keys:
            difficulty, stage = split_stage_key(key)
            items = list(self._data[key])
            for idx, (sec, dt, _notice) in enumerate(items, 1):
                if not isinstance(dt, datetime):
                    dt = _parse_dt(dt)
                ws_hist.cell(row_i, 1, difficulty)
                ws_hist.cell(row_i, 2, stage)
                ws_hist.cell(row_i, 3, sec)
                ws_hist.cell(row_i, 4, dt.strftime("%Y-%m-%d %H:%M:%S"))
                ws_hist.cell(row_i, 5, idx)
                row_i += 1
        for col in range(1, 6):
            ws_hist.column_dimensions[get_column_letter(col)].width = 18
        return wb

    def _write_workbook_safe(self) -> tuple:
        """写入 Excel。主文件被占用时写入 clear_times_最新.xlsx。

        返回 (ok, path, error_message)
        """
        wb = self._build_workbook()
        candidates = [
            self.path,
            self.path.parent / "clear_times_最新.xlsx",
        ]
        last_err = ""
        tmp = self.path.parent / f".~clear_times_{os.getpid()}.xlsx"
        try:
            wb.save(tmp)
        finally:
            wb.close()

        for target in candidates:
            try:
                if target.exists():
                    try:
                        shutil.copy2(target, target.with_suffix(target.suffix + ".bak"))
                    except Exception:
                        pass
                mid = target.with_suffix(target.suffix + ".partial")
                shutil.copy2(tmp, mid)
                mid.replace(target)
                try:
                    if tmp.exists():
                        tmp.unlink()
                except Exception:
                    pass
                return True, target, ""
            except Exception as exc:
                last_err = str(exc)
                continue
        try:
            if tmp.exists():
                tmp.unlink()
        except Exception:
            pass
        return False, self.path, last_err or "excel write failed"

    def _write_workbook(self) -> None:
        """兼容旧调用：失败抛错。"""
        ok, _path, err = self._write_workbook_safe()
        if not ok:
            raise OSError(err or "excel write failed")

    def excel_path(self) -> Path:
        return self.path

    def json_store_path(self) -> Path:
        return self.json_path
